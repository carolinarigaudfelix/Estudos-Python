# openpyxl para arquivos Excel xlsx, xlsm, xltx e xltm (instalação)
# Com essa biblioteca será possível ler e escrever dados em células
# específicas, formatar células, inserir gráficos,
# criar fórmulas, adicionar imagens e outros elementos gráficos às suas
# planilhas. Ela é útil para automatizar tarefas envolvendo planilhas do
# Excel, como a criação de relatórios e análise de dados e/ou facilitando a
# manipulação de grandes quantidades de informações.
# Instalação necessária: pip install openpyxl
# Documentação: https://openpyxl.readthedocs.io/en/stable/


from pathlib import Path
from openpyxl import Workbook
from openpyxl.worksheet.worksheet import Worksheet #tipando worksheet

ROOT_FOLDER = Path(__file__).parent
WORKBOOK_PATH = ROOT_FOLDER / 'workbook.xlsx'

workbook = Workbook()
#worksheet: Worksheet = workbook.active


#criar uma planilha 
#nome para a planilha
sheet_name = 'Mina planilha'
#cria a planilha
workbook.create_sheet(sheet_name, 0) #põe no índice 0 como planilha ativa
#selecinoa a planilha
worksheet: Worksheet = workbook[sheet_name]
print(workbook.sheetnames)

#Remover sheet
workbook.remove(workbook['Sheet']) #remove planilha

#  woorksheet.  agora funciona
worksheet.cell(1, 1 , 'Nome') #L1 C1 nome
worksheet.cell(1, 2 , 'Idade') 
worksheet.cell(1, 3 , 'Nota') 


#Criando os cabeçalhos


students = [
    #nome           idade nota
    ['João',           14,  3.2],
    ['Maria',           13,  5.5],
    ['Luiz',           15,  8.8],
    ['Sofia',           18,  9.5],

]

# for i in range(2, 10):      #Logica de coluna linha
#     for j in range(1, 4):
#         print(f'Linha {i}, Coluna {j}')

# for i, students_row in enumerate(students, start=2):
#     for j, students_column in enumerate(students_row, start=1):
#         worksheet.cell(i, j, students_column)

for student in students:
    worksheet.append(student)

workbook.save(WORKBOOK_PATH)