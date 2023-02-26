# openpyxl - ler e alterar dados de um planilha
# específicas, formatar células, inserir gráficos,
# criar fórmulas, adicionar imagens e outros elementos gráficos às suas
# planilhas. Ela é útil para automatizar tarefas envolvendo planilhas do
# Excel, como a criação de relatórios e análise de dados e/ou facilitando a
# manipulação de grandes quantidades de informações.
# Instalação necessária: pip install openpyxl
# Documentação: https://openpyxl.readthedocs.io/en/stable/


from pathlib import Path
from openpyxl.cell import Cell
from openpyxl import Workbook, load_workbook
from openpyxl.worksheet.worksheet import Worksheet #tipando worksheet

ROOT_FOLDER = Path(__file__).parent
WORKBOOK_PATH = ROOT_FOLDER / 'workbook.xlsx'

#carregando um arquivo do excel
workbook: Workbook = load_workbook(WORKBOOK_PATH)

#Nome para a planilha
sheet_name = 'Mina planilha'

#selecionando  a planilha
worksheet: Worksheet = workbook[sheet_name]

row: tuple[Cell]
for row in worksheet.iter_rows():
    for cell in row:
        print(cell.value, end ='\t')

        if cell.value =='Maria':
            # worksheet['B3'].value=23 #altera
            worksheet.cell(cell.row, 2, 23) #cell.row considera Maria, 2 é a col
    print() #printa o valor


workbook.save(WORKBOOK_PATH)